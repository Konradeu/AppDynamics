import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import yaml
from functions import getAppID, getBtRules, getBTs, getTiers, elemntComparison, applicationCheck, fillingList, getEumRules, getEumApps, fillAndConcatenateDicts

import asyncio


'''The Script Itself'''

async def main():


    # Reading the configuration.yaml file
    with open('configuration.yaml', 'r') as file:
        configs = yaml.safe_load(file)

        prodSecret = configs['Prod']['api_client_secret']
        prodClient = configs['Prod']['api_client_username']
        prodURL = configs['Prod']['controller_url']
        prodAccount = configs['Prod']['controller_account']
        prodApp = configs['Prod']['app_name']
        prodEumApp = configs['Prod']['eum_app_name']

        stagingSecret = configs['Staging']['api_client_secret']
        stagingClient = configs['Staging']['api_client_username']
        stagingURL = configs['Staging']['controller_url']
        stagingAccount = configs['Staging']['controller_account']
        stagingApp = configs['Staging']['app_name']
        stagingEumApp = configs['Staging']['eum_app_name']
        file.close()

    # Authentication - requesting the Bearer Token to make API Calls to the Prod Controller
    auth_credentials = {"grant_type": "client_credentials", "client_id": prodClient + "@" + prodAccount, "client_secret": prodSecret}
    auth_response = requests.post(prodURL + "/controller/api/oauth/access_token", data=auth_credentials, headers={"Content-Type": "application/x-www-form-urlencoded"})
    print("Authentication (for Bearer Token) response: " + str(auth_response.status_code))

    prodAuth_token = auth_response.json()["access_token"]
    prodApplicationNameList, prodApplicationStatusList = await applicationCheck(prodURL, prodAuth_token)


    # Getting the Prod Data
    prodAppID = getAppID(prodURL, prodApp, prodAuth_token)
    prodRuleNamesList, prodRuleEnabledList = getBtRules(prodURL, prodAppID, prodAuth_token)

    prodDataDict = await  getBTs(prodURL, prodApp, prodAuth_token)
    # getBtMetric(prodURL, prodApp, prodDataDict['Tier'][0], prodDataDict['BT names'][0], 120, prodAuth_token)

    prodTiersNameList, prodTiersStatusList = getTiers(prodURL, prodApp, prodAuth_token)

    # Authentication - requesting the Bearer Token to make API Calls to the Test Controller
    auth_credentials = {"grant_type": "client_credentials", "client_id": stagingClient + "@" + stagingAccount, "client_secret": stagingSecret}
    auth_response = requests.post(stagingURL + "/controller/api/oauth/access_token", data=auth_credentials, headers={"Content-Type": "application/x-www-form-urlencoded"})
    print("Authentication (for Bearer Token) response: " + str(auth_response.status_code))

    stagingAuth_token = auth_response.json()["access_token"]

    stagingApplicationNameList, stagingApplicationStatusList = await applicationCheck(stagingURL, stagingAuth_token)

    # Getting the Staging Data
    stagingAppID = getAppID(stagingURL, stagingApp, stagingAuth_token)
    stagingRuleNamesList, stagingRuleEnabledList = getBtRules(stagingURL, stagingAppID, stagingAuth_token)
    stagingDataDict = await getBTs(stagingURL, stagingApp, stagingAuth_token)
    stagingTiersNameList, stagingTiersStatusList = getTiers(stagingURL, stagingApp, stagingAuth_token)


    # Logic for checking whether the BT's from Staging are present in the Prod
    isInListList = []

    for btName in stagingDataDict['BT names']:
        if btName in prodDataDict['BT names']: isInListList.append('Yes')
        else: isInListList.append('No')

    # Creating PD Dataframes for saving the data to a Excxel file
    btProdDt = pd.DataFrame.from_dict(prodDataDict)
    btCompDt = pd.DataFrame.from_dict({'BT names': stagingDataDict['BT names'], 'Has load?': stagingDataDict['Has load?'], 'IsInProd?': isInListList, 'Used Rule': stagingDataDict['Used Rule'], 'Entry Point': stagingDataDict['Entry Point'], 'Tier': stagingDataDict['Tier']})

    prodRuleNamesList, prodRuleEnabledList, stagingRuleNamesList, stagingRuleEnabledList, rulesIsInBothEvents = elemntComparison(prodRuleNamesList, prodRuleEnabledList, stagingRuleNamesList, stagingRuleEnabledList)

    prodTiersNameList, prodTiersStatusList, stagingTiersNameList, stagingTiersStatusList, tiersIsInBothEvents = elemntComparison(prodTiersNameList, prodTiersStatusList, stagingTiersNameList, stagingTiersStatusList)

    btRulesDt = pd.DataFrame.from_dict({'Prod': prodRuleNamesList, 'Prod_Enabled': prodRuleEnabledList, 'Staging': stagingRuleNamesList, 'Staging_Enabled': stagingRuleEnabledList, 'IsInBothEnvs?': rulesIsInBothEvents})

    prodApplicationNameList, prodApplicationStatusList, stagingApplicationNameList, stagingApplicationStatusList, appsIsInBothEvents  = elemntComparison(prodApplicationNameList, prodApplicationStatusList, stagingApplicationNameList, stagingApplicationStatusList)
    applicationsDf = pd.DataFrame.from_dict({'Prod': prodApplicationNameList, 'Prod_Load': prodApplicationStatusList, 'Staging': stagingApplicationNameList, 'Staging_Load': stagingApplicationStatusList, "IsInBothEnvs": appsIsInBothEvents})

    # Tier PD

    tiersDf = pd.DataFrame.from_dict({'Prod': prodTiersNameList, 'Prod_Status': prodTiersStatusList, 'Staging': stagingTiersNameList, 'Staging_Status': stagingTiersStatusList, 'IsInBothEnvs': tiersIsInBothEvents})

    # SummaryR
    sumamryDict = {"Number of Prod BTs": [len(prodDataDict['BT names'])], 'Number of BTs with Load': [prodDataDict['Has load?'].count('Yes')], "Number of Staging BTs": [len(stagingDataDict['BT names'])], 'Number of BTs with Load': [stagingDataDict['Has load?'].count('Yes')], 'BT Limit': [200]}
    summaryPd = pd.DataFrame.from_dict(sumamryDict)

    # EUM Part

    prodEumApps = getEumApps(prodURL, prodAuth_token)
    stagingEumApps = getEumApps(stagingURL, stagingAuth_token)

    prodIterator, prodEumRulesDict = getEumRules(prodURL, prodAuth_token, prodEumApp, prodEumApps, 'Prod')
    stagingIterator, stagingEumRulesDict = getEumRules(stagingURL, stagingAuth_token, stagingEumApp, stagingEumApps, 'Staging')

    prodEumAppsNames, prodEumAppsActive, stagingEumAppsNames, stagingEumAppsActive = fillingList([e.name for e in prodEumApps], [e.active for e in prodEumApps], [e.name for e in stagingEumApps], [e.active for e in stagingEumApps])

    eumAppsDf = pd.DataFrame.from_dict({'Prod': prodEumAppsNames, 'ProdActive': prodEumAppsActive, 'Staging': stagingEumAppsNames , 'StagingActive': stagingEumAppsActive})

    EumBase_to_df = fillAndConcatenateDicts(prodEumRulesDict['Includes'], stagingEumRulesDict['Includes'])

    eumIncludeRulesDf = pd.DataFrame.from_dict(EumBase_to_df)

    EumBase_to_df = fillAndConcatenateDicts(prodEumRulesDict['Excludes'], stagingEumRulesDict['Excludes'])

    eumExcludeRulesDf = pd.DataFrame.from_dict(EumBase_to_df)


    # Part where the non-existing BT indicator cells will get red background colour
    with pd.ExcelWriter('BT_Prd_Staging_Comparator/BT_Comparison_Result.xlsx') as writer:
        btProdDt.to_excel(writer, sheet_name="Prod")
        btCompDt.to_excel(writer, sheet_name="Staging")
        btRulesDt.to_excel(writer, sheet_name="Rules")
        tiersDf.to_excel(writer, sheet_name="Tiers")
        applicationsDf.to_excel(writer, sheet_name="Appplications")
        eumAppsDf.to_excel(writer, sheet_name="EUMAppplications")

        eumIncludeRulesDf.to_excel(writer, sheet_name="IncludeBasePages")

        eumExcludeRulesDf.to_excel(writer, sheet_name="ExcludeBasePages")


    wb = openpyxl.load_workbook("BT_Prd_Staging_Comparator/BT_Comparison_Result.xlsx") #path to the Excel file
    ws = wb['Staging'] #Name of the working sheet

    fill_cell = PatternFill(patternType='solid',
                                fgColor='C64747') #You can give the hex code for different color
    ws.fill = fill_cell # B2 is the name of the cell to be fill with color

    for cells in ws['D']:
        if cells.value == 'No': cells.fill = PatternFill(patternType='solid', fgColor='00FF0000')

    wb.save("BT_Comparison_Result.xlsx")

if __name__ == '__main__':
    asyncio.run(main())